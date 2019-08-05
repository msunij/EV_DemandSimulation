import openpyxl
import pickle

wb = openpyxl.load_workbook('connection_properties_zeroed.xlsx')
sheet = wb.active


prop_dict = {}

#Number of connection
pk_weekend_count = [ sheet.cell(row=23,column=10).value, sheet.cell(row=23,column=11).value]
pk_weekday_count = [ sheet.cell(row=24,column=10).value, sheet.cell(row=24,column=11).value]

#Connection time

pk_weekday_1 = []
pk_weekday_2 = []
pk_weekend_1 = []
pk_weekend_2 = []

for i in range(6,102):
	pk_weekday_1.append(sheet.cell(row=i, column=4).value)
	pk_weekday_2.append(sheet.cell(row=i, column=5).value)
	pk_weekend_1.append(sheet.cell(row=i, column=6).value)
	pk_weekend_2.append(sheet.cell(row=i, column=7).value)


#State of charge (SOC)
pk_weekend_1_initial = []
pk_weekend_2_initial = []
pk_weekday_1_initial = []
pk_weekday_2_initial = []
pk_weekend_1_final = []
pk_weekend_2_final = []
pk_weekday_1_final = []
pk_weekday_2_final = []

for i in range(5,18):
	pk_weekday_1_initial.append(sheet.cell(row=i, column=10).value)
	pk_weekday_2_initial.append(sheet.cell(row=i, column=11).value)
	pk_weekday_1_final.append(sheet.cell(row=i, column=12).value)
	pk_weekday_2_final.append(sheet.cell(row=i, column=13).value)
	pk_weekend_1_initial.append(sheet.cell(row=i, column=14).value)
	pk_weekend_2_initial.append(sheet.cell(row=i, column=15).value)
	pk_weekend_1_final.append(sheet.cell(row=i, column=16).value)
	pk_weekend_2_final.append(sheet.cell(row=i, column=17).value)

# Normalizing the probabilites
pk_weekend_count = [ p/sum(pk_weekend_count) for p in pk_weekend_count ]
pk_weekday_count = [ p/sum(pk_weekday_count) for p in pk_weekday_count ]

pk_weekday_1 = [ p/sum(pk_weekday_1) for p in pk_weekday_1 ]
pk_weekday_2 = [ p/sum(pk_weekday_2) for p in pk_weekday_2 ]
pk_weekend_1 = [ p/sum(pk_weekend_1) for p in pk_weekend_1 ]
pk_weekend_2 = [ p/sum(pk_weekend_2) for p in pk_weekend_2 ]

pk_weekend_1_initial = [ p/sum(pk_weekend_1_initial) for p in pk_weekend_1_initial]
pk_weekend_2_initial = [ p/sum(pk_weekend_2_initial) for p in pk_weekend_2_initial]
pk_weekday_1_initial = [ p/sum(pk_weekday_1_initial) for p in pk_weekday_1_initial]
pk_weekday_2_initial = [ p/sum(pk_weekday_2_initial) for p in pk_weekday_2_initial]
pk_weekend_1_final = [ p/sum(pk_weekend_1_final) for p in pk_weekend_1_final]
pk_weekend_2_final = [ p/sum(pk_weekend_2_final) for p in pk_weekend_2_final]
pk_weekday_1_final = [ p/sum(pk_weekday_1_final) for p in pk_weekday_1_final]
pk_weekday_2_final = [ p/sum(pk_weekday_2_final) for p in pk_weekday_2_final]

#time step and soc level
timestep_connection = [ i for i in range(0,1440,15)]
soc_unit_level = [ i for i in range(13)]

#add it to the dictionaray
prob_dict = {'count_prob_day':pk_weekday_count,
							'count_prob_end': pk_weekend_count,
							'connection_time_step':timestep_connection,
							'connection_time_day_1':pk_weekday_1, 
							'connection_time_day_2':pk_weekday_2, 
							'connection_time_end_1':pk_weekend_1, 
							'connection_time_end_2':pk_weekend_2,
							'soc_unit_level': soc_unit_level, 
							'soc_day_initial_1':pk_weekday_1_initial, 
							'soc_day_initial_2':pk_weekday_2_initial, 
							'soc_day_final_1':pk_weekday_1_final, 
							'soc_day_final_2':pk_weekday_2_final, 
							'soc_end_initial_1':pk_weekend_1_initial,  
							'soc_end_initial_2':pk_weekend_2_initial,  
							'soc_end_final_1':pk_weekend_1_final,  
							'soc_end_final_2':pk_weekend_2_final}

#pickle dump process
with open('normalized_probabilities_zeroed.pickle','wb') as fp:
	pickle.dump(prob_dict,fp)